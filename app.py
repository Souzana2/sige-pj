import streamlit as st
import pandas as pd
import subprocess
import sys
import time
import datetime
from conect import get_connection

st.set_page_config(page_title="SIGE Admin", page_icon="⚙️", layout="wide")
st.title("⚙️ SIGE - Painel de Administração")
st.divider()

def atualizar_excel_stock_produto(produto_id, nova_qtd, novas_vendas, novo_stock_min, produto_nome=None):
    """Atualiza a quantidade_atual, vendas_mensais e stock_minimo do produto no Excel dados.xlsx em tempo real."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook("dados.xlsx")
        ws = wb["stock"]
        col_id, col_prod = None, None
        col_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "id": col_id = col
            if val == "produto": col_prod = col
            if val in ["quantidade_atual", "vendas_mensais", "stock_minimo"]:
                col_map[val] = col
                
        linha_alvo = None
        if col_id:
            for row in range(2, ws.max_row + 1):
                cell_id = ws.cell(row=row, column=col_id).value
                if cell_id is not None and int(cell_id) == int(produto_id):
                    linha_alvo = row
                    break
        if linha_alvo is None and col_prod and produto_nome:
            for row in range(2, ws.max_row + 1):
                cell_prod = ws.cell(row=row, column=col_prod).value
                if cell_prod is not None and str(cell_prod).strip() == str(produto_nome).strip():
                    linha_alvo = row
                    break
                    
        if linha_alvo:
            if "quantidade_atual" in col_map:
                ws.cell(row=linha_alvo, column=col_map["quantidade_atual"]).value = int(nova_qtd)
            if "vendas_mensais" in col_map:
                ws.cell(row=linha_alvo, column=col_map["vendas_mensais"]).value = int(novas_vendas)
            if "stock_minimo" in col_map:
                ws.cell(row=linha_alvo, column=col_map["stock_minimo"]).value = int(novo_stock_min)
                
        wb.save("dados.xlsx")
        wb.close()
    except Exception as e:
        st.warning(f"Aviso: Não foi possível atualizar o Excel 'dados.xlsx'. Detalhe: {e}")

def atualizar_excel_ativo_funcionario(funcionario_nome, ativo_val):
    """Atualiza a coluna ativo (1/0) do funcionário no Excel dados.xlsx."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook("dados.xlsx")
        ws = wb["funcionarios"]
        col_nome, col_ativo = None, None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "nome": col_nome = col
            if val == "ativo": col_ativo = col
        if col_nome and col_ativo:
            for row in range(2, ws.max_row + 1):
                cell_nome = ws.cell(row=row, column=col_nome).value
                if cell_nome is not None and str(cell_nome).strip() == str(funcionario_nome).strip():
                    ws.cell(row=row, column=col_ativo).value = int(ativo_val)
                    break
        wb.save("dados.xlsx")
        wb.close()
    except Exception as e:
        print(f"Erro ao atualizar ativo do funcionário no Excel: {e}")

def atualizar_excel_ativo_produto(produto_nome, ativo_val):
    """Atualiza a coluna ativo (1/0) do produto no Excel dados.xlsx."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook("dados.xlsx")
        ws = wb["stock"]
        col_prod, col_ativo = None, None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "produto": col_prod = col
            if val == "ativo": col_ativo = col
        if col_prod and col_ativo:
            for row in range(2, ws.max_row + 1):
                cell_prod = ws.cell(row=row, column=col_prod).value
                if cell_prod is not None and str(cell_prod).strip() == str(produto_nome).strip():
                    ws.cell(row=row, column=col_ativo).value = int(ativo_val)
                    break
        wb.save("dados.xlsx")
        wb.close()
    except Exception as e:
        print(f"Erro ao atualizar ativo do produto no Excel: {e}")

# ==========================================
# FUNÇÃO DE CARREGAMENTO
# ==========================================

def carregar_funcionarios():
    conn = get_connection()
    # 💡 Repare aqui: Adicionei o fin.custo_turnover_previsto mesmo antes do risco
    query = """
    SELECT f.id, f.nome, f.departamento, f.cargo, f.salario, 
           fin.custo_total_mensal, fin.custo_turnover_previsto, f.risco_saida
    FROM funcionarios f
    LEFT JOIN financeiro_funcionarios fin ON f.id = fin.id_funcionario
    WHERE f.ativo = 1
    """
    df = pd.read_sql(query, conn)
    conn.close()
    
    if 'risco_saida' in df.columns:
        df['risco_saida'] = pd.to_numeric(df['risco_saida'], errors='coerce').fillna(0).round(0).astype(int)
        
    df['Desligar'] = False 
    return df


def carregar_stock():
    conn = get_connection()
    # 💡 Repare aqui: Adicionei a previsao_ruptura no final
    query = """
    SELECT id, produto, categoria, quantidade_atual, vendas_mensais, 
           preco_compra, preco_venda, previsao_ruptura
    FROM stock
    WHERE ativo = 1
    """
    df = pd.read_sql(query, conn)
    conn.close()
    
    if 'previsao_ruptura' in df.columns:
        df['previsao_ruptura'] = pd.to_numeric(df['previsao_ruptura'], errors='coerce').fillna(0).round(0).astype(int)
        
    df['Desligar'] = False 
    return df

tab_func, tab_stock, tab_dash = st.tabs(["👥 Funcionários", "📦 Stock", "📈 Dashboard Financeiro"])
with tab_func:
    
    # ==========================================
    # 1º ANDAR: REGISTAR (Mantido limpo como gostou)
    # ==========================================
    with st.expander("➕ Registar Novo Funcionário"):
        with st.form("form_novo_func", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                nome = st.text_input("Nome Completo *")
                data_nascimento = st.date_input(
                    "Data de Nascimento",
                    min_value=datetime.date(1930, 1, 1),
                    max_value=datetime.date.today(),
                    value=datetime.date(1995, 1, 1)
                )
                data_admissao = st.date_input("Data de Admissão *", format="DD/MM/YYYY")
            with col2:
                departamento = st.text_input("Departamento")
                cargo = st.text_input("Cargo")
                salario = st.number_input("Salário Mensal (€) *", min_value=0.0, step=50.0)
                horas_extra = st.number_input("Horas Extra (mensais)", min_value=0.0, step=1.0, value=0.0)

            if st.form_submit_button("✅ Guardar Novo Funcionário"):
                if nome == "" or salario <= 0:
                    st.warning("⚠️ O Nome e o Salário são obrigatórios!")
                else:
                    _func_ok = False
                    try:
                        conn = get_connection()
                        cursor = conn.cursor()

                        # Inicia com percentagens padrão de 100%
                        assiduidade          = 100.0
                        produtividade        = 100.0
                        satisfacao           = 100.0
                        avaliacao_desempenho = 100.0

                        sql_func = """INSERT INTO funcionarios
                                      (nome, data_nascimento, data_admissao, departamento, cargo,
                                       salario, horas_extra, assiduidade, produtividade, satisfacao,
                                       avaliacao_desempenho, ativo)
                                      VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)"""
                        cursor.execute(sql_func, (nome, data_nascimento, data_admissao, departamento, cargo,
                                                  salario, horas_extra, assiduidade, produtividade, satisfacao,
                                                  avaliacao_desempenho))

                        novo_id = cursor.lastrowid
                        valor_hora = salario / 160
                        custo_he   = horas_extra * valor_hora * 1.5
                        encargos   = round(salario * 0.2375, 2)
                        custo_total = round(salario + encargos + custo_he, 2)
                        turnover    = round(salario * 2, 2)

                        sql_fin = """INSERT INTO financeiro_funcionarios
                                     (id_funcionario, salario, custo_horas_extra, encargos,
                                      custo_total_mensal, custo_turnover_previsto)
                                     VALUES (%s, %s, %s, %s, %s, %s)"""
                        cursor.execute(sql_fin, (novo_id, salario, custo_he, encargos, custo_total, turnover))
                        conn.commit()
                        _func_ok = True
                    except Exception as e:
                        st.error(f"Erro: {e}")
                    finally:
                        if 'cursor' in locals(): cursor.close()
                        if 'conn' in locals(): conn.close()
                    if _func_ok:
                        st.success("✅ Funcionário registado com sucesso!")
                        time.sleep(1.2)
                        st.rerun()

    # ==========================================
    # NOVO ANDAR: REGISTAR PAGAMENTO DE FUNCIONÁRIO
    # ==========================================
    with st.expander("💸 Registar Pagamento / Lançamento de Funcionário"):
        try:
            conn = get_connection()
            df_funcs_lista = pd.read_sql("SELECT id, nome, salario FROM funcionarios WHERE ativo = 1", conn)
            conn.close()
        except Exception as e:
            df_funcs_lista = pd.DataFrame()
            st.error(f"Erro ao carregar lista de funcionários: {e}")

        if not df_funcs_lista.empty:
            lista_f = df_funcs_lista['id'].astype(str) + " - " + df_funcs_lista['nome']
            
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                func_sel = st.selectbox("Selecione o Funcionário:", lista_f, key="sb_pay_func")
            with col_f2:
                tipo_pag_sel = st.selectbox("Tipo de Lançamento/Pagamento:", ["Salário", "Bónus", "Adiantamento", "Outro"])
                
            id_func_pay = int(func_sel.split(" - ")[0])
            func_row = df_funcs_lista[df_funcs_lista['id'] == id_func_pay].iloc[0]
            salario_base = float(func_row['salario'] or 0.0)
            
            # Sugere o salário base se for selecionado Salário
            valor_sugerido = salario_base if tipo_pag_sel == "Salário" else 0.0
            
            with st.form("form_novo_pagamento_funcionario", clear_on_submit=True):
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    valor_pag = st.number_input("Valor Pago (€) *", min_value=0.0, step=10.0, value=valor_sugerido)
                    obs_pag = st.text_input("Observação (Opcional)", placeholder="ex: Referente ao mês de Junho/2026")
                with col_p2:
                    data_pag = st.date_input("Data do Pagamento *", value=datetime.date.today())
                
                if st.form_submit_button("💾 Gravar Pagamento"):
                    if valor_pag <= 0:
                        st.error("❌ O valor do pagamento deve ser maior que 0.")
                    else:
                        _pay_ok = False
                        try:
                            conn = get_connection()
                            cursor = conn.cursor()
                            sql_ins_pay = """INSERT INTO pagamentos_funcionarios 
                                             (id_funcionario, tipo_pagamento, valor, data_pagamento, observacao)
                                             VALUES (%s, %s, %s, %s, %s)"""
                            cursor.execute(sql_ins_pay, (id_func_pay, tipo_pag_sel, valor_pag, data_pag, obs_pag))
                            conn.commit()
                            _pay_ok = True
                        except Exception as e:
                            if 'conn' in locals(): conn.rollback()
                            st.error(f"Erro ao registar pagamento: {e}")
                        finally:
                            if 'cursor' in locals(): cursor.close()
                            if 'conn' in locals(): conn.close()
                            
                        if _pay_ok:
                            st.success(f"✅ Pagamento de {tipo_pag_sel} no valor de {valor_pag:,.2f} € para '{func_row['nome']}' registado com sucesso!")
                            time.sleep(1.2)
                            st.rerun()
        else:
            st.info("Registe primeiro um funcionário ativo para poder registar pagamentos.")

    # ==========================================
    # NOVO ANDAR: AVALIAÇÕES E NOTAS OBTIDAS
    # ==========================================
    st.subheader("📊 Notas e Avaliações Obtidas")
    try:
        conn = get_connection()
        df_av = pd.read_sql("""
            SELECT nome, assiduidade, produtividade, satisfacao, avaliacao_desempenho, risco_saida
            FROM funcionarios WHERE ativo = 1
        """, conn)
        conn.close()
        
        if not df_av.empty:
            df_av_display = pd.DataFrame()
            df_av_display["Nome do Funcionário"] = df_av["nome"]
            df_av_display["Assiduidade"] = df_av["assiduidade"].map(lambda x: f"{x:.0f}%" if pd.notna(x) else "100%")
            df_av_display["Produtividade"] = df_av["produtividade"].map(lambda x: f"{x:.0f}%" if pd.notna(x) else "100%")
            df_av_display["Satisfação"] = df_av["satisfacao"].map(lambda x: f"{x:.0f}%" if pd.notna(x) else "100%")
            df_av_display["Avaliação Desempenho"] = df_av["avaliacao_desempenho"].map(lambda x: f"{x:.0f}%" if pd.notna(x) else "100%")
            
            def map_estado(risco):
                r = float(risco or 0)
                if r >= 75:
                    return "🔴 Em Risco"
                elif r >= 40:
                    return "🟡 Avaliar"
                else:
                    return "🟢 Seguro"
            
            df_av_display["Estado do Risco"] = df_av["risco_saida"].apply(map_estado)
            
            st.dataframe(df_av_display, use_container_width=True, hide_index=True)
        else:
            st.info("Sem funcionários ativos registados.")
    except Exception as e:
        st.error(f"Erro ao carregar avaliações: {e}")

    st.write("---")
    
    # ==========================================
    # 2º ANDAR: PESQUISA E TABELA (Só de Leitura e Desligar)
    # ==========================================
    
    st.subheader("🔍 Tabela Resumo da Equipa")

    # ==========================================
    # BOTÃO DA INTELIGÊNCIA ARTIFICIAL
    # ==========================================
    if st.button("🧠 Rodar Inteligência Artificial (Recalcular Riscos)", type="primary", use_container_width=True):
        with st.spinner("A Inteligência Artificial está a analisar o comportamento da equipa... Aguarde."):
            try:
                # O Python vai rodar o seu ficheiro de ML nos bastidores!
                # Certifique-se de que o nome do ficheiro aqui está IGUAL ao que você tem no VS Code
                result = subprocess.run(
                    [sys.executable, "ml_funcionarios.py"],
                    capture_output=True, text=True
                )
                if result.returncode == 0:
                    st.success("✅ Previsões concluídas! A base de dados foi atualizada pela IA.")
                else:
                    st.warning(f"⚠️ ML terminou com avisos:\n{result.stderr[-300:] if result.stderr else ''}")
                time.sleep(1.5)
                st.rerun()

            except subprocess.CalledProcessError:
                st.error("❌ O ficheiro de ML deu um erro. Verifique se o código no 'ml_funcionarios.py' está a rodar bem sozinho.")
            except FileNotFoundError:
                st.error("❌ Ficheiro 'ml_funcionarios.py' não encontrado. Verifique o nome!")
    
    df_func = carregar_funcionarios()
    termo_pesquisa = st.text_input("Buscar funcionário globalmente (Nome, Cargo, etc):")
    
    if termo_pesquisa:
        filtro_global = (
            df_func['nome'].astype(str).str.contains(termo_pesquisa, case=False, na=False) |
            df_func['departamento'].astype(str).str.contains(termo_pesquisa, case=False, na=False) |
            df_func['cargo'].astype(str).str.contains(termo_pesquisa, case=False, na=False)
        )
        df_func = df_func[filtro_global]
    
    # 💡 TRANCAR TUDO EXCETO O "DESLIGAR"
    colunas_trancadas = [col for col in df_func.columns if col != 'Desligar']

    
    df_func_editado = st.data_editor(
        df_func,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        disabled=colunas_trancadas,
        key="edit_func",
        column_config={
            "id": st.column_config.NumberColumn("ID", width="small"),
            "nome": st.column_config.TextColumn("Nome"),
            "departamento": st.column_config.TextColumn("Departamento"),
            "cargo": st.column_config.TextColumn("Cargo"),
            "salario": st.column_config.NumberColumn(
                "Salário (€)",
                format="%.2f €",
                width="small",
            ),
            "custo_total_mensal": st.column_config.NumberColumn(
                "Custo Total/Mês (€)",
                format="%.2f €",
                help="Salário + Encargos Patronais (23,75%) + Horas Extra (×1,5)",
            ),
            "custo_turnover_previsto": st.column_config.NumberColumn(
                "Custo Turnover Previsto (€)",
                format="%.2f €",
                help="Custo estimado de substituição: equivalente a 2× salário mensal bruto",
            ),
            "risco_saida": st.column_config.ProgressColumn(
                "Risco de Saída (%)",
                min_value=0,
                max_value=100,
                format="%d%%",
                help="Índice de risco calculado pela IA (0 = seguro, 100 = crítico)",
            ),
            "Desligar": st.column_config.CheckboxColumn("Desligar?", width="small"),
        },
    )
    
    if st.button("💾 Executar Desligamentos Selecionados", type="primary"):
        conn = get_connection()
        cursor = conn.cursor()
        for index, row in df_func_editado.iterrows():
            if row['Desligar'] == True:
                cursor.execute("UPDATE funcionarios SET ativo = 0 WHERE id = %s", (int(row['id']),))
                atualizar_excel_ativo_funcionario(row['nome'], 0)
        conn.commit()
        st.success("✅ Guardado com sucesso!")
        time.sleep(1.2) # 💡 O ecrã congela por 1.2 segundos para a pessoa ler a mensagem
        st.rerun()

    st.write("---")

    # 3º ANDAR: EDIÇÃO COMPLETA
    st.subheader("✏️ Ficha de Edição Completa")
    lista_funcionarios = df_func['id'].astype(str) + " - " + df_func['nome']
    
    # Caixa de seleção simples (sem o key=session_state)
    escolha = st.selectbox("Escolha quem deseja editar:", ["(Selecione um funcionário)"] + lista_funcionarios.tolist())

    # Só abre a janela se alguém estiver selecionado
    if escolha != "(Selecione um funcionário)":
        id_selecionado = int(escolha.split(" - ")[0])
        
        # Vai ao SQL buscar TUDO o que existe sobre esta pessoa
        conn = get_connection()
        df_perfil = pd.read_sql(f"SELECT * FROM funcionarios WHERE id = {id_selecionado}", conn)
        conn.close()
        
        if not df_perfil.empty:
            dados = df_perfil.iloc[0]
            
            with st.form("form_edicao_completa"):
                st.write(f"A editar ficha completa de: **{dados['nome']}**")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    novo_nome = st.text_input("Nome", value=dados['nome'])
                    # Atenção: O Python precisa de converter a data que vem do MySQL para mostrar no calendário
                    # Forma clássica, limpa e sem erros de sintaxe!
                    try:
                        if pd.notna(dados['data_nascimento']):
                            data_nasc_atual = pd.to_datetime(dados['data_nascimento']).date()
                        else:
                            data_nasc_atual = datetime.date(1995, 1, 1)
                    except:
                        data_nasc_atual = datetime.date(1995, 1, 1)
                    
                    nova_data_nasc = st.date_input(
                        "Data de Nasc.", 
                        value=data_nasc_atual,
                        min_value=datetime.date(1930, 1, 1), 
                        max_value=datetime.date.today()
                    )
                    nova_data_adm = st.date_input("Data Admissão", value=dados['data_admissao'])
                with col2:
                    novo_dep    = st.text_input("Departamento", value=dados['departamento'] if dados['departamento'] else "")
                    novo_cargo  = st.text_input("Cargo",        value=dados['cargo'] if dados['cargo'] else "")
                    novo_salario = st.number_input("Salário", value=float(dados['salario']), step=50.0)
                    novas_horas  = st.number_input("Horas Extra", min_value=0.0, value=float(dados['horas_extra'] or 0))
                with col3:
                    st.caption("📋 Avaliação de Desempenho")

                    # Percentagens actuais (para pré-calcular os campos brutos)
                    _ass = float(dados.get('assiduidade') or 100)
                    _pro = float(dados.get('produtividade') or 100)
                    _sat = float(dados.get('satisfacao') or 100)
                    _ava = float(dados.get('avaliacao_desempenho') or 100)
                    _ris = float(dados.get('risco_saida') or 0)
                    alerta_emoji = "🔴" if _ris >= 75 else ("🟡" if _ris >= 40 else "🟢")
                    alerta_txt   = "ALTO" if _ris >= 75 else ("MÉDIO" if _ris >= 40 else "BAIXO")
                    st.caption(f"{alerta_emoji} Risco de Saída (IA): **{_ris:.0f}%** — {alerta_txt}")

                    # Valores brutos inferidos a partir dos % actuais
                    _dias = 22
                    ed_dias   = st.number_input("Dias Úteis no Mês",      min_value=1, max_value=31, step=1, value=_dias)
                    ed_faltas = st.number_input("Faltas",                  min_value=0, max_value=31, step=1,
                                                value=max(0, round((1 - _ass/100) * _dias)))
                    ed_m_atr  = st.number_input("Metas Atribuídas",        min_value=1, step=1, value=10)
                    ed_m_con  = st.number_input("Metas Concluídas",        min_value=0, step=1,
                                                value=max(0, min(10, round(_pro/100 * 10))))
                    ed_nota   = st.select_slider("Nota Satisfação (1-5)",  options=[1,2,3,4,5],
                                                 value=max(1, min(5, round(_sat/100 * 5))))
                    ed_gest   = st.select_slider("Avaliação Gestor (1-5)", options=[1,2,3,4,5],
                                                 value=max(1, min(5, round(_ava/100 * 5))))

                if st.form_submit_button("💾 Guardar Todas as Alterações", type="primary"):
                    conn = get_connection()
                    cursor = conn.cursor()

                    # Calcula percentagens a partir dos dados brutos
                    nova_assiduidade   = round(max(0.0, (ed_dias - ed_faltas) / ed_dias * 100), 2) if ed_dias > 0 else 100.0
                    nova_produtividade = round(ed_m_con / ed_m_atr * 100, 2) if ed_m_atr > 0 else 100.0
                    nova_satisfacao    = round(ed_nota / 5 * 100, 2)
                    nova_avaliacao     = round(ed_gest / 5 * 100, 2)

                    sql_update = """UPDATE funcionarios
                                    SET nome=%s, data_nascimento=%s, data_admissao=%s, departamento=%s, cargo=%s,
                                        salario=%s, horas_extra=%s, assiduidade=%s, produtividade=%s,
                                        satisfacao=%s, avaliacao_desempenho=%s
                                    WHERE id=%s"""
                    cursor.execute(sql_update, (novo_nome, nova_data_nasc, nova_data_adm, novo_dep, novo_cargo,
                                                novo_salario, novas_horas, nova_assiduidade, nova_produtividade,
                                                nova_satisfacao, nova_avaliacao, id_selecionado))

                    # Recalcula Finanças
                    valor_hora  = novo_salario / 160
                    custo_he    = novas_horas * valor_hora * 1.5
                    encargos    = round(novo_salario * 0.2375, 2)
                    custo_total = round(novo_salario + encargos + custo_he, 2)
                    turnover    = round(novo_salario * 2, 2)

                    sql_fin = """UPDATE financeiro_funcionarios
                                 SET salario=%s, custo_horas_extra=%s, encargos=%s,
                                     custo_total_mensal=%s, custo_turnover_previsto=%s
                                 WHERE id_funcionario=%s"""
                    cursor.execute(sql_fin, (novo_salario, custo_he, encargos, custo_total, turnover, id_selecionado))

                    conn.commit()
                    st.session_state.id_selecionado_edit = "(Selecione um funcionário)"
                    st.success("✅ Guardado com sucesso!")
                    time.sleep(1.2)
                    st.rerun()

# --- ABA: STOCK ---
with tab_stock:
    
    # ==========================================
    # 1º ANDAR: REGISTAR NOVO PRODUTO
    # ==========================================
    with st.expander("➕ Registar Novo Produto"):

        with st.form("form_novo_stock", clear_on_submit=True):
            st.write("Introduza as características do novo produto. Quantidades, vendas e mínimos iniciam a 0 e são calculados de forma automática.")
            col1, col2 = st.columns(2)
            with col1:
                produto      = st.text_input("Nome do Produto *")
                categoria    = st.text_input("Categoria")
                tempo_reposicao = st.number_input("Tempo Reposição (dias) *", min_value=0, step=1, value=5)
            with col2:
                preco_compra = st.number_input("Preço Compra (€)", min_value=0.0, step=0.01, value=0.0)
                preco_venda  = st.number_input("Preço Venda (€)",  min_value=0.0, step=0.01, value=0.0)

            if st.form_submit_button("✅ Adicionar ao Armazém"):
                if produto == "":
                    st.warning("⚠️ O Nome do Produto é obrigatório!")
                else:
                    _stock_ok = False
                    try:
                        conn = get_connection()
                        cursor = conn.cursor()
                        sql_stock = """INSERT INTO stock
                                       (produto, categoria, quantidade_atual, vendas_mensais,
                                        reposicoes, preco_compra, preco_venda,
                                        stock_minimo, tempo_reposicao, ativo)
                                       VALUES (%s, %s, 0, 0, 0, %s, %s, 0, %s, 1)"""
                        cursor.execute(sql_stock, (produto, categoria, preco_compra, preco_venda, tempo_reposicao))
                        novo_prod_id = cursor.lastrowid

                        # Calcula e insere financeiro automaticamente se houver preços (vendas_mensais = 0)
                        if preco_compra > 0 and preco_venda > 0:
                            cursor.execute(
                                """INSERT INTO financeiro
                                   (id_produto, preco_compra, preco_venda, vendas_mensais,
                                    faturamento_mensal, custo_mensal, lucro_mensal,
                                    margem_lucro, giro_capital)
                                   VALUES (%s, %s, %s, 0, 0.0, 0.0, 0.0, 0.0, 0.0)""",
                                (novo_prod_id, preco_compra, preco_venda)
                            )

                        conn.commit()
                        _stock_ok = True
                    except Exception as e:
                        if 'conn' in locals(): conn.rollback()
                        st.error(f"Erro ao registar: {e}")
                    finally:
                        if 'cursor' in locals(): cursor.close()
                        if 'conn' in locals(): conn.close()
                    if _stock_ok:
                        st.success("✅ Produto adicionado ao armazém com sucesso!")
                        time.sleep(1.2)
                        st.rerun()

    # ==========================================
    # NOVO ANDAR: REGISTAR TRANSAÇÃO (ENTRADA / SAÍDA)
    # ==========================================
    with st.expander("💸 Registar Entrada / Saída de Stock (Transações)"):
        try:
            conn = get_connection()
            df_prods = pd.read_sql("SELECT id, produto, quantidade_atual, preco_compra, preco_venda FROM stock WHERE ativo = 1", conn)
            conn.close()
        except Exception as e:
            df_prods = pd.DataFrame()
            st.error(f"Erro ao carregar lista de produtos: {e}")

        if not df_prods.empty:
            lista_prods = df_prods['id'].astype(str) + " - " + df_prods['produto']
            
            col_sel1, col_sel2 = st.columns(2)
            with col_sel1:
                prod_sel = st.selectbox("Selecione o Produto para Movimentar:", lista_prods, key="sb_mov_prod")
            with col_sel2:
                tipo_mov_sel = st.radio("Tipo de Movimento:", ["Entrada (Compra de Stock)", "Saída (Venda de Stock)"], horizontal=True)
                
            id_prod_mov = int(prod_sel.split(" - ")[0])
            prod_row = df_prods[df_prods['id'] == id_prod_mov].iloc[0]
            qtd_atual_prod = int(prod_row['quantidade_atual'])
            preco_compra_padrao = float(prod_row['preco_compra'] or 0.0)
            preco_venda_padrao = float(prod_row['preco_venda'] or 0.0)
            
            is_saida = "Saída" in tipo_mov_sel
            preco_sugerido = preco_venda_padrao if is_saida else preco_compra_padrao
            
            with st.form("form_nova_movimentacao", clear_on_submit=True):
                col_m1, col_m2 = st.columns(2)
                with col_m1:
                    qtd_mov = st.number_input("Quantidade (unid.)", min_value=1, step=1, value=1)
                    obs_mov = st.text_input("Observação (Opcional)", placeholder="ex: Venda de balcão, Compra a fornecedor")
                with col_m2:
                    preco_mov = st.number_input("Preço Unitário (€)", min_value=0.0, step=0.01, value=preco_sugerido)
                
                st.caption(f"Stock Atual em Armazém: **{qtd_atual_prod}** unidades.")
                
                if st.form_submit_button("💾 Gravar Movimentação"):
                    if is_saida and qtd_mov > qtd_atual_prod:
                        st.error(f"❌ Stock insuficiente! Tentou vender {qtd_mov} unidades, mas só existem {qtd_atual_prod} em armazém.")
                    else:
                        _mov_ok = False
                        try:
                            # 1. Calcular nova quantidade
                            if is_saida:
                                nova_qtd = qtd_atual_prod - qtd_mov
                                db_tipo = "Saída"
                            else:
                                nova_qtd = qtd_atual_prod + qtd_mov
                                db_tipo = "Entrada"
                                
                            conn = get_connection()
                            cursor = conn.cursor()
                            
                            # 2. Gravar na tabela movimentacoes
                            sql_ins_mov = """INSERT INTO movimentacoes 
                                             (id_produto, tipo_movimento, quantidade, preco_unitario, observacao)
                                             VALUES (%s, %s, %s, %s, %s)"""
                            cursor.execute(sql_ins_mov, (id_prod_mov, db_tipo, qtd_mov, preco_mov, obs_mov))
                            
                            # 3. Calcular automático de vendas_mensais (últimos 30 dias) e stock_minimo
                            # (incluindo a transação recém-inserida)
                            cursor.execute("SELECT tempo_reposicao FROM stock WHERE id = %s", (id_prod_mov,))
                            tempo_rep = int(cursor.fetchone()[0] or 0)
                            
                            query_vendas_30d = """
                                SELECT SUM(quantidade) 
                                FROM movimentacoes 
                                WHERE id_produto = %s AND tipo_movimento = 'Saída' AND data_hora >= NOW() - INTERVAL 30 DAY
                            """
                            cursor.execute(query_vendas_30d, (id_prod_mov,))
                            vendas_30d = int(cursor.fetchone()[0] or 0)
                            
                            import math
                            vendas_diarias = vendas_30d / 30.0
                            stock_min = int(math.ceil(vendas_diarias * tempo_rep))
                            
                            # 4. Atualizar na tabela stock
                            sql_upd_stock = """UPDATE stock 
                                               SET quantidade_atual = %s, vendas_mensais = %s, stock_minimo = %s 
                                               WHERE id = %s"""
                            cursor.execute(sql_upd_stock, (nova_qtd, vendas_30d, stock_min, id_prod_mov))
                            
                            # 5. Atualizar tabela financeiro para o produto para refletir as novas vendas reais
                            fat_m = round(preco_venda_padrao * vendas_30d, 2)
                            cus_m = round(preco_compra_padrao * vendas_30d, 2)
                            luc_m = round(fat_m - cus_m, 2)
                            mar_l = round((luc_m / fat_m) * 100, 4) if fat_m > 0 else 0.0
                            gir_c = round(fat_m / preco_compra_padrao, 4) if preco_compra_padrao > 0 else 0.0
                            
                            sql_upd_fin = """
                                INSERT INTO financeiro
                                (id_produto, preco_compra, preco_venda, vendas_mensais,
                                 faturamento_mensal, custo_mensal, lucro_mensal,
                                 margem_lucro, giro_capital)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                preco_compra=%s, preco_venda=%s, vendas_mensais=%s,
                                faturamento_mensal=%s, custo_mensal=%s, lucro_mensal=%s,
                                margem_lucro=%s, giro_capital=%s
                            """
                            cursor.execute(sql_upd_fin, (
                                id_prod_mov, preco_compra_padrao, preco_venda_padrao, vendas_30d,
                                fat_m, cus_m, luc_m, mar_l, gir_c,
                                preco_compra_padrao, preco_venda_padrao, vendas_30d,
                                fat_m, cus_m, luc_m, mar_l, gir_c
                            ))
                            
                            conn.commit()
                            _mov_ok = True
                            
                            # 6. Atualizar no Excel dados.xlsx (quantidade, vendas e stock_minimo)
                            atualizar_excel_stock_produto(id_prod_mov, nova_qtd, vendas_30d, stock_min, produto_nome=prod_row['produto'])
                            
                        except Exception as e:
                            if 'conn' in locals(): conn.rollback()
                            st.error(f"Erro ao registar movimentação: {e}")
                        finally:
                            if 'cursor' in locals(): cursor.close()
                            if 'conn' in locals(): conn.close()
                            
                        if _mov_ok:
                            st.success(f"✅ Movimentação de {db_tipo} de {qtd_mov} unidades de '{prod_row['produto']}' registada com sucesso!")
                            time.sleep(1.2)
                            st.rerun()
        else:
            st.info("Registe primeiro um produto no armazém para poder movimentar stock.")

    # ==========================================================
    # NOVO ANDAR: ALERTAS DE RUPTURA DE STOCK
    # ==========================================
    st.subheader("📊 Alertas de Ruptura de Stock")
    try:
        conn = get_connection()
        df_stock_av = pd.read_sql("""
            SELECT produto, quantidade_atual, vendas_mensais, tempo_reposicao, previsao_ruptura
            FROM stock WHERE ativo = 1
        """, conn)
        conn.close()
        
        if not df_stock_av.empty:
            # Calcula os dias até ruptura
            def calc_dias(row):
                vendas = float(row['vendas_mensais'] or 0)
                qtd = float(row['quantidade_atual'] or 0)
                lead_time = float(row['tempo_reposicao'] or 0)
                
                vendas_diarias = vendas / 30.0
                if vendas_diarias > 0:
                    return int(round((qtd / vendas_diarias) - lead_time))
                else:
                    return 999
            
            df_stock_av["Dias até Ruptura"] = df_stock_av.apply(calc_dias, axis=1)
            
            # Estado do Risco baseado nos dias de margem vs. tempo de reposição do produto
            # <= 0   : ruptura iminente ou já ocorreu              → Em Risco
            # 0..lead: stock esgota antes da próxima encomenda     → Avaliar
            # > lead : há margem suficiente                        → Seguro
            def map_estado_dias(row):
                dias = row["Dias até Ruptura"]
                lead = float(row["tempo_reposicao"] or 20)
                if dias <= 0:
                    return "🔴 Em Risco"
                elif dias <= lead:
                    return "🟡 Avaliar"
                else:
                    return "🟢 Seguro"

            df_stock_av["Estado do Risco"] = df_stock_av.apply(map_estado_dias, axis=1)
            
            # Ordena: ruptura primeiro, depois por dias ascendente
            df_stock_av = df_stock_av.sort_values(by=["previsao_ruptura", "Dias até Ruptura"], ascending=[False, True])
            
            # Seleciona apenas as colunas pedidas
            df_stock_display = df_stock_av[["produto", "Dias até Ruptura", "Estado do Risco"]].copy()
            df_stock_display.columns = ["Nome do Produto", "Dias até Ruptura", "Estado do Risco"]
            
            # Coloração vermelha para dias negativos usando Pandas Styler
            def color_dias_negativos(val):
                if isinstance(val, (int, float)) and val < 0:
                    return 'color: red; font-weight: bold;'
                return ''
            
            st.dataframe(
                df_stock_display.style.map(color_dias_negativos, subset=["Dias até Ruptura"]),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("Sem produtos ativos no inventário.")
    except Exception as e:
        st.error(f"Erro ao carregar alertas de stock: {e}")

    st.write("---")
    
    # ==========================================
    # 2º ANDAR: PESQUISA E TABELA RESUMO
    # ==========================================
    st.subheader("📦 Visão Geral do Armazém")
    # ==========================================
    # BOTÃO DA INTELIGÊNCIA ARTIFICIAL (STOCK)
    # ==========================================
    if st.button("🧠 Rodar Inteligência Artificial (Prever Rupturas)", type="primary", use_container_width=True, key="btn_ml_stock"):
        with st.spinner("A Inteligência Artificial está a analisar as vendas e quantidades no armazém... Aguarde."):
            try:
                # ⚠️ ATENÇÃO AQUI: Substitua "ml_stock.py" pelo nome exato do seu ficheiro Python de IA do Stock!
                result_ml = subprocess.run(
                    [sys.executable, "ml_stock.py"],
                    capture_output=True, text=True
                )
                if result_ml.returncode == 0:
                    st.success("✅ Previsões de Ruptura atualizadas com sucesso na base de dados!")
                else:
                    st.warning(f"⚠️ ML Stock terminou com avisos:\n{result_ml.stderr[-300:] if result_ml.stderr else ''}")
                time.sleep(1.5)
                st.rerun()

            except subprocess.CalledProcessError:
                st.error("❌ O modelo de IA do Stock deu um erro interno. Verifique se ele roda bem sozinho.")
            except FileNotFoundError:
                st.error("❌ Ficheiro não encontrado! Escreveu o nome correto dentro do comando subprocess.run?")
    
    try:
        df_stock = carregar_stock()
        termo_pesquisa_stock = st.text_input("Buscar produto (Nome ou Categoria):")
        
        if termo_pesquisa_stock:
            filtro_stock = (
                df_stock['produto'].astype(str).str.contains(termo_pesquisa_stock, case=False, na=False) |
                df_stock['categoria'].astype(str).str.contains(termo_pesquisa_stock, case=False, na=False)
            )
            df_stock = df_stock[filtro_stock]
        
        colunas_trancadas_stock = [col for col in df_stock.columns if col != 'Desligar']
        
        df_stock_editado = st.data_editor(
            df_stock,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=colunas_trancadas_stock,
            key="edit_stock",
            column_config={
                "id": st.column_config.NumberColumn("ID", width="small"),
                "produto": st.column_config.TextColumn("Produto"),
                "categoria": st.column_config.TextColumn("Categoria"),
                "quantidade_atual": st.column_config.NumberColumn(
                    "Qtd. em Armazém",
                    format="%d un.",
                    help="Unidades fisicamente em stock neste momento",
                    width="small",
                ),
                "vendas_mensais": st.column_config.NumberColumn(
                    "Vendas/Mês",
                    format="%d un.",
                    help="Unidades vendidas por mês (usado para calcular dias até ruptura)",
                    width="small",
                ),
                "preco_compra": st.column_config.NumberColumn(
                    "Preço Compra (€)",
                    format="%.2f €",
                    help="Custo unitário de aquisição ao fornecedor",
                ),
                "preco_venda": st.column_config.NumberColumn(
                    "Preço Venda (€)",
                    format="%.2f €",
                    help="Preço de venda ao cliente",
                ),
                "previsao_ruptura": st.column_config.ProgressColumn(
                    "Risco de Ruptura (%)",
                    min_value=0,
                    max_value=100,
                    format="%d%%",
                    help="0% = stock confortável · 100% = ruptura iminente (calculado pela IA)",
                ),
                "Desligar": st.column_config.CheckboxColumn("Remover?", width="small"),
            },
        )
        
        if st.button("💾 Executar Remoções de Catálogo", type="primary"):
            conn = get_connection()
            cursor = conn.cursor()
            for index, row in df_stock_editado.iterrows():
                if row['Desligar'] == True:
                    cursor.execute("UPDATE stock SET ativo = 0 WHERE id = %s", (int(row['id']),))
                    atualizar_excel_ativo_produto(row['produto'], 0)
            conn.commit()
            st.success("✅ Guardado com sucesso!")
            time.sleep(1.2) # 💡 O ecrã congela por 1.2 segundos para a pessoa ler a mensagem
            st.rerun()
            
    except Exception as e:
        st.error(f"Erro na tabela: {e}")

    st.write("---")

    # ==========================================
    # 3º ANDAR: EDIÇÃO COMPLETA DE PRODUTO
    # ==========================================
    st.subheader("✏️ Atualizar Ficha do Produto")
    
    try:
        lista_produtos = df_stock['id'].astype(str) + " - " + df_stock['produto']
        escolha_prod = st.selectbox("Escolha o produto para atualizar:", ["(Selecione um produto)"] + lista_produtos.tolist())

        if escolha_prod != "(Selecione um produto)":
            id_prod_selecionado = int(escolha_prod.split(" - ")[0])

            # Vai buscar todos os campos do produto directamente à BD
            conn_ed = get_connection()
            df_perfil_prod = pd.read_sql(
                f"SELECT * FROM stock WHERE id = {id_prod_selecionado}", conn_ed
            )
            conn_ed.close()

            if not df_perfil_prod.empty:
                dados_prod = df_perfil_prod.iloc[0]

                with st.form("form_edicao_stock"):
                    st.write(f"A atualizar: **{dados_prod['produto']}**")

                    colA, colB = st.columns(2)
                    with colA:
                        novo_nome_prod = st.text_input("Nome", value=str(dados_prod['produto']))
                        nova_cat       = st.text_input("Categoria", value=str(dados_prod['categoria']) if pd.notna(dados_prod['categoria']) else "")
                        novo_tempo_rep  = st.number_input("Tempo Reposição (dias)", min_value=0, step=1, value=int(dados_prod['tempo_reposicao'] or 0))
                    with colB:
                        novo_preco_c   = st.number_input("Preço Compra (€)", min_value=0.0, step=0.01, value=float(dados_prod['preco_compra'] or 0))
                        novo_preco_v   = st.number_input("Preço Venda (€)",  min_value=0.0, step=0.01, value=float(dados_prod['preco_venda'] or 0))

                    if st.form_submit_button("💾 Guardar e Atualizar Stock", type="primary"):
                        _edit_ok = False
                        try:
                            conn = get_connection()
                            cursor = conn.cursor()
                            sql_upd_stock = """UPDATE stock
                                               SET produto=%s, categoria=%s,
                                                   preco_compra=%s, preco_venda=%s,
                                                   tempo_reposicao=%s
                                               WHERE id=%s"""
                            cursor.execute(sql_upd_stock, (
                                novo_nome_prod, nova_cat,
                                novo_preco_c, novo_preco_v,
                                novo_tempo_rep,
                                id_prod_selecionado
                            ))

                            # Recalcula e sincroniza a tabela financeiro usando vendas atuais da BD
                            vm    = int(dados_prod.get('vendas_mensais') or 0)
                            fat_m = round(novo_preco_v * vm, 2)
                            cus_m = round(novo_preco_c * vm, 2)
                            luc_m = round(fat_m - cus_m, 2)
                            mar_l = round((luc_m / fat_m) * 100, 4) if fat_m > 0 else 0.0
                            gir_c = round(fat_m / novo_preco_c, 4) if novo_preco_c > 0 else 0.0
                            cursor.execute(
                                """INSERT INTO financeiro
                                   (id_produto, preco_compra, preco_venda, vendas_mensais,
                                    faturamento_mensal, custo_mensal, lucro_mensal,
                                    margem_lucro, giro_capital)
                                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                   ON DUPLICATE KEY UPDATE
                                   preco_compra=%s, preco_venda=%s, vendas_mensais=%s,
                                   faturamento_mensal=%s, custo_mensal=%s, lucro_mensal=%s,
                                   margem_lucro=%s, giro_capital=%s""",
                                (id_prod_selecionado, novo_preco_c, novo_preco_v, vm,
                                 fat_m, cus_m, luc_m, mar_l, gir_c,
                                 novo_preco_c, novo_preco_v, vm,
                                 fat_m, cus_m, luc_m, mar_l, gir_c)
                            )
                            conn.commit()
                            _edit_ok = True
                            
                            # Atualiza no Excel dados.xlsx
                            atualizar_excel_stock_produto(
                                id_prod_selecionado, 
                                dados_prod['quantidade_atual'], 
                                vm, 
                                dados_prod['stock_minimo'],
                                produto_nome=dados_prod['produto']
                            )
                        except Exception as e:
                            if 'conn' in locals(): conn.rollback()
                            st.error(f"Erro ao guardar: {e}")
                        finally:
                            if 'cursor' in locals(): cursor.close()
                            if 'conn' in locals(): conn.close()
                        if _edit_ok:
                            st.success("✅ Guardado com sucesso!")
                            time.sleep(1.2)
                            st.rerun()
    except Exception as e:
        st.error(f"Erro ao processar edição: {e}")

# ==========================================
# --- ABA: DASHBOARD FINANCEIRO ---
# ==========================================
with tab_dash:

    # -----------------------------------------------------------------------
    # Formatador português: 1.234.567,89 €
    # -----------------------------------------------------------------------
    def fmt_eur(value, decimals=2):
        formatted = f"{value:,.{decimals}f}"
        # Troca separadores: , → X (temp) → . ; . → ,
        formatted = formatted.replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"{formatted} €"
    # -----------------------------------------------------------------------
    
    try:
        # 1. Carregar as duas tabelas
        df_f = carregar_funcionarios()
        df_s = carregar_stock()
        
        # 2. Limpar valores nulos para as contas não darem erro
        df_s.fillna(0, inplace=True)
        df_f.fillna(0, inplace=True)
        
        # 3. A MATEMÁTICA FINANCEIRA
        # -- Receitas Estimadas --
        faturacao_mensal = (df_s['vendas_mensais'] * df_s['preco_venda']).sum()
        custo_mercadoria = (df_s['vendas_mensais'] * df_s['preco_compra']).sum()
        custo_pessoal = df_f['custo_total_mensal'].sum()
        receita_por_funcionario = faturacao_mensal / len(df_f) if len(df_f) > 0 else 0
        capital_empatado = (df_s['quantidade_atual'] * df_s['preco_compra']).sum()
        risco_turnover_total = df_f['custo_turnover_previsto'].sum()

        # -- Receitas e Custos Reais (Movimentações do Mês Atual) --
        hoje = datetime.date.today()
        primeiro_dia_mes = hoje.replace(day=1)
        
        try:
            conn = get_connection()
            # Saídas (vendas) do mês atual
            query_saidas = """
                SELECT SUM(quantidade * preco_unitario) as faturacao_real 
                FROM movimentacoes 
                WHERE tipo_movimento = 'Saída' AND data_hora >= %s
            """
            cursor = conn.cursor()
            cursor.execute(query_saidas, (primeiro_dia_mes,))
            faturacao_real = float(cursor.fetchone()[0] or 0.0)
            
            # Entradas (compras) do mês atual
            query_entradas = """
                SELECT SUM(quantidade * preco_unitario) as custo_mercadoria_real 
                FROM movimentacoes 
                WHERE tipo_movimento = 'Entrada' AND data_hora >= %s
            """
            cursor.execute(query_entradas, (primeiro_dia_mes,))
            custo_mercadoria_real = float(cursor.fetchone()[0] or 0.0)

            # Pagamentos de pessoal reais (salários, bónus, etc.) do mês atual
            query_pagamentos = """
                SELECT SUM(valor) as custo_pessoal_real 
                FROM pagamentos_funcionarios 
                WHERE data_pagamento >= %s
            """
            cursor.execute(query_pagamentos, (primeiro_dia_mes,))
            custo_pessoal_real = float(cursor.fetchone()[0] or 0.0)
            
            # Top 3 produtos reais vendidos no mês (Curva ABC Real)
            query_top_reais = """
                SELECT s.produto, SUM(m.quantidade * m.preco_unitario) as Receita_Mensal
                FROM movimentacoes m
                JOIN stock s ON m.id_produto = s.id
                WHERE m.tipo_movimento = 'Saída' AND m.data_hora >= %s
                GROUP BY s.produto
                ORDER BY Receita_Mensal DESC
                LIMIT 3
            """
            df_top_reais = pd.read_sql(query_top_reais, conn, params=(primeiro_dia_mes,))
            cursor.close()
            conn.close()
        except Exception as db_e:
            faturacao_real = 0.0
            custo_mercadoria_real = 0.0
            custo_pessoal_real = 0.0
            df_top_reais = pd.DataFrame(columns=["produto", "Receita_Mensal"])
            st.error(f"Erro ao buscar movimentações reais da BD: {db_e}")

        # 4. DESENHAR OS CARTÕES (KPIs)
        st.subheader("📊 Indicadores de Saúde Financeira")
        
        # Filtro/Toggle de perspetiva
        visao = st.radio(
            "Selecione a perspetiva financeira:", 
            ["Visão Real (Baseada em Transações do Mês)", "Visão Estimada (Baseada em Projeções de Vendas)"], 
            horizontal=True
        )
        
        if visao == "Visão Real (Baseada em Transações do Mês)":
            fat_display = faturacao_real
            custo_merc_display = custo_mercadoria_real
            custo_pes_display = custo_pessoal_real
            custo_tot_display = custo_mercadoria_real + custo_pessoal_real
            lucro_display = fat_display - custo_tot_display
            legenda_fat = "Faturação Real (Vendas do Mês)"
            legenda_custo_mat = f"Pessoal Real: {fmt_eur(custo_pes_display, 0)} | Compras Real: {fmt_eur(custo_merc_display, 0)}"
            legenda_lucro = "✅ Lucro Real" if lucro_display >= 0 else "🚨 Prejuízo Real"
            receita_pc = fat_display / len(df_f) if len(df_f) > 0 else 0
            top_display = df_top_reais.copy()
            if not top_display.empty:
                top_display["Receita_Mensal"] = top_display["Receita_Mensal"].apply(lambda v: fmt_eur(v, 2))
            st.info("ℹ️ **Visão Real:** inclui o valor total das compras de stock efetuadas este mês como custo (reposições de armazém). "
                    "Meses com grande reabastecimento podem mostrar prejuízo mesmo com vendas saudáveis. "
                    "Compare com a **Visão Estimada** para ver a margem operacional projetada.")
        else:
            fat_display = faturacao_mensal
            custo_merc_display = custo_mercadoria
            custo_pes_display = custo_pessoal
            custo_tot_display = custo_mercadoria + custo_pessoal
            lucro_display = faturacao_mensal - custo_tot_display
            legenda_fat = "Faturação Estimada (Projeções)"
            legenda_custo_mat = f"Pessoal Estimado: {fmt_eur(custo_pes_display, 0)} | Matéria Estimada: {fmt_eur(custo_merc_display, 0)}"
            legenda_lucro = "✅ Lucro Estimado" if lucro_display >= 0 else "🚨 Prejuízo Estimado"
            receita_pc = receita_por_funcionario
            
            df_s['Receita_Mensal'] = df_s['vendas_mensais'] * df_s['preco_venda']
            top_display = df_s.nlargest(3, 'Receita_Mensal')[['produto', 'Receita_Mensal']].copy()
            top_display["Receita_Mensal"] = top_display["Receita_Mensal"].apply(lambda v: fmt_eur(v, 2))

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Faturação do Mês", fmt_eur(fat_display), help=legenda_fat)
        col2.metric("Custos Operacionais", fmt_eur(custo_tot_display), legenda_custo_mat, delta_color="inverse")
        col3.metric("Resultado Líquido", fmt_eur(lucro_display), legenda_lucro)
        col4.metric("Receita per Capita", fmt_eur(receita_pc), "Faturação gerada por empregado")
        
        st.write("---")

        # 5. GRÁFICOS E DETALHES
        colA, colB = st.columns(2)
        
        with colA:
            st.subheader("💰 Distribuição de Custos Reais")
            # REMOVEMOS O TURNOVER DAQUI! Fica apenas o dinheiro que sai da conta.
            df_grafico = pd.DataFrame({
                "Categoria": ["Pessoal", "Mercadoria (Compras/Matéria)"],
                "Valor (€)": [custo_pes_display, custo_merc_display]
            }).set_index("Categoria")
            
            # Gráfico de barras simples
            st.bar_chart(df_grafico, color="#ff4b4b")
            
            # Novo bloco de Alerta de Risco para o Turnover
            st.warning(f"⚠️ **Risco Latente (Turnover):** Se os funcionários atualmente em risco de saída abandonarem a empresa, estima-se um impacto oculto de **{fmt_eur(risco_turnover_total)}** (Indemnizações, novas contratações e quebra de produtividade).")
            
        with colB:
            st.subheader("📦 Giro de Capital")
            st.metric("Capital Empatado em Armazém", fmt_eur(capital_empatado), "Dinheiro investido no Stock atual", delta_color="off")
            
            st.write(f"**Top 3 Produtos que mais faturam ({'Vendas Reais' if visao.startswith('Visão Real') else 'Projeções'}):**")
            if not top_display.empty:
                st.dataframe(top_display, hide_index=True, use_container_width=True)
            else:
                st.info("Sem registos de vendas para este período.")

    except Exception as e:
        st.error(f"⚠️ Ainda não há dados suficientes para gerar o Dashboard ou ocorreu um erro: {e}")