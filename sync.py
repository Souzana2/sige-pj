import pandas as pd
from openpyxl import load_workbook
from conect import get_connection

print("Iniciando a Sincronizacao Inteligente...")

# ==========================================
# FOLHAS DO EXCEL QUE PERTENCEM AO PIPELINE
# Qualquer outra folha (ex: guias) e ignorada
# ==========================================
FOLHAS_SYNC = ["funcionarios", "financeiro_funcionarios", "stock", "financeiro"]

# ==========================================
# FUNÇÕES DE LIMPEZA
# ==========================================
def tratar_dinheiro(valor):
    """Limpa euros e converte para numero que o SQL entende."""
    if pd.isna(valor) or valor == "": return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    try:
        return float(str(valor).replace('€', '').replace(' ', '').replace('.', '').replace(',', '.'))
    except:
        return 0.0

def tratar_nulo(valor, padrao=None):
    """Transforma celulas vazias em NULL para o SQL."""
    if pd.isna(valor) or valor == "": return padrao
    return valor

def tratar_data(valor):
    """Converte datas para o padrao do SQL (AAAA-MM-DD)."""
    if pd.isna(valor) or valor == "": return None
    try: return pd.to_datetime(valor, dayfirst=True).strftime('%Y-%m-%d')
    except: return None

def tratar_percentagem(valor):
    """
    Converte percentagens do Excel para escala 0-100.
    O Excel guarda '60%' como 0.6 internamente.
    Se o valor for <= 1.0 assume que e fraccao e multiplica por 100.
    """
    if pd.isna(valor) or valor == "": return None
    v = float(valor)
    if 0.0 <= v <= 1.0:
        return round(v * 100, 4)
    return round(v, 4)

# ==========================================
# SINCRONIZAÇÃO: FUNCIONÁRIOS
# Devolve dict {linha_excel: id_gerado} para linhas novas
# ==========================================
def sincronizar_funcionarios(cursor, df):
    print("-> A processar Funcionarios...")
    df = df.dropna(subset=['nome'])
    id_writes = {}   # linhas novas que precisam de ID escrito no Excel

    for index, row in df.iterrows():
        linha_excel = index + 2
        salario     = tratar_dinheiro(row.get("salario"))
        horas_extra = tratar_dinheiro(row.get("horas_extra", row.get("hora_extra")))

        assiduidade          = tratar_percentagem(row.get("assiduidade"))
        produtividade        = tratar_percentagem(row.get("produtividade"))
        satisfacao           = tratar_percentagem(row.get("satisfacao"))
        avaliacao_desempenho = tratar_percentagem(row.get("avaliacao_desempenho"))

        tem_id = not (pd.isna(row.get("id")) or row.get("id") == "")

        if tem_id:
            id_val = int(row["id"])
            cursor.execute(
                """UPDATE funcionarios
                   SET nome=%s, data_nascimento=%s, departamento=%s, cargo=%s,
                       data_admissao=%s, salario=%s, horas_extra=%s,
                       assiduidade=%s, produtividade=%s, satisfacao=%s,
                       avaliacao_desempenho=%s, ativo=%s
                   WHERE id=%s""",
                (
                    tratar_nulo(row.get("nome")), tratar_data(row.get("data_nascimento")),
                    tratar_nulo(row.get("departamento")), tratar_nulo(row.get("cargo")),
                    tratar_data(row.get("data_admissao")), salario, horas_extra,
                    assiduidade, produtividade, satisfacao, avaliacao_desempenho,
                    tratar_nulo(row.get("ativo"), 1), id_val
                )
            )
            if cursor.rowcount == 0:
                cursor.execute(
                    """INSERT IGNORE INTO funcionarios
                       (id, nome, data_nascimento, departamento, cargo,
                        data_admissao, salario, horas_extra,
                        assiduidade, produtividade, satisfacao,
                        avaliacao_desempenho, ativo)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        id_val,
                        tratar_nulo(row.get("nome")), tratar_data(row.get("data_nascimento")),
                        tratar_nulo(row.get("departamento")), tratar_nulo(row.get("cargo")),
                        tratar_data(row.get("data_admissao")), salario, horas_extra,
                        assiduidade, produtividade, satisfacao, avaliacao_desempenho,
                        tratar_nulo(row.get("ativo"), 1)
                    )
                )
        else:
            cursor.execute(
                """INSERT INTO funcionarios
                   (nome, data_nascimento, departamento, cargo, data_admissao,
                    salario, horas_extra, assiduidade, produtividade, satisfacao,
                    avaliacao_desempenho, ativo)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    tratar_nulo(row.get("nome")), tratar_data(row.get("data_nascimento")),
                    tratar_nulo(row.get("departamento")), tratar_nulo(row.get("cargo")),
                    tratar_data(row.get("data_admissao")), salario, horas_extra,
                    assiduidade, produtividade, satisfacao, avaliacao_desempenho,
                    tratar_nulo(row.get("ativo"), 1)
                )
            )
            id_writes[linha_excel] = cursor.lastrowid  # ID gerado pelo MySQL

    return id_writes


def sincronizar_financeiro_funcionarios(cursor, df):
    print("-> A processar Financeiro dos Funcionarios...")
    df = df.dropna(subset=['id_funcionario'])

    for _, row in df.iterrows():
        id_val = int(row["id_funcionario"])
        sal = tratar_dinheiro(row.get("salario"))
        che = tratar_dinheiro(row.get("custo_horas_extra"))
        enc = tratar_dinheiro(row.get("encargos"))
        tot = tratar_dinheiro(row.get("custo_total_mensal"))
        ctp = tratar_dinheiro(row.get("custo_turnover_previsto"))

        cursor.execute("SELECT COUNT(*) FROM financeiro_funcionarios WHERE id_funcionario=%s", (id_val,))
        existe = cursor.fetchone()[0]

        if existe == 0:
            cursor.execute(
                """INSERT INTO financeiro_funcionarios
                   (id_funcionario, salario, custo_horas_extra, encargos,
                    custo_total_mensal, custo_turnover_previsto)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (id_val, sal, che, enc, tot, ctp)
            )
        else:
            cursor.execute(
                """UPDATE financeiro_funcionarios
                   SET salario=%s, custo_horas_extra=%s, encargos=%s,
                       custo_total_mensal=%s, custo_turnover_previsto=%s
                   WHERE id_funcionario=%s""",
                (sal, che, enc, tot, ctp, id_val)
            )


# ==========================================
# SINCRONIZAÇÃO: STOCK
# Devolve dict {linha_excel: id_gerado} para linhas novas
# ==========================================
def sincronizar_stock(cursor, df):
    print("-> A processar Stock...")
    df = df.dropna(subset=['produto'])
    id_writes = {}

    for index, row in df.iterrows():
        linha_excel = index + 2
        tem_id = not (pd.isna(row.get("id")) or row.get("id") == "")

        if tem_id:
            id_val = int(row["id"])
            cursor.execute(
                """UPDATE stock
                   SET produto=%s, categoria=%s, quantidade_atual=%s, vendas_mensais=%s,
                       reposicoes=%s, preco_compra=%s, preco_venda=%s,
                       stock_minimo=%s, tempo_reposicao=%s
                   WHERE id=%s""",
                (
                    tratar_nulo(row.get("produto")), tratar_nulo(row.get("categoria")),
                    tratar_nulo(row.get("quantidade_atual"), 0), tratar_nulo(row.get("vendas_mensais"), 0),
                    tratar_nulo(row.get("reposicoes"), 0), tratar_dinheiro(row.get("preco_compra")),
                    tratar_dinheiro(row.get("preco_venda")), tratar_nulo(row.get("stock_minimo"), 0),
                    tratar_nulo(row.get("tempo_reposicao"), 0), id_val
                )
            )
            if cursor.rowcount == 0:
                cursor.execute(
                    """INSERT IGNORE INTO stock
                       (id, produto, categoria, quantidade_atual, vendas_mensais,
                        reposicoes, preco_compra, preco_venda,
                        stock_minimo, tempo_reposicao, previsao_ruptura)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        id_val,
                        tratar_nulo(row.get("produto")), tratar_nulo(row.get("categoria")),
                        tratar_nulo(row.get("quantidade_atual"), 0), tratar_nulo(row.get("vendas_mensais"), 0),
                        tratar_nulo(row.get("reposicoes"), 0), tratar_dinheiro(row.get("preco_compra")),
                        tratar_dinheiro(row.get("preco_venda")), tratar_nulo(row.get("stock_minimo"), 0),
                        tratar_nulo(row.get("tempo_reposicao"), 0), tratar_nulo(row.get("previsao_ruptura"))
                    )
                )
        else:
            cursor.execute(
                """INSERT INTO stock
                   (produto, categoria, quantidade_atual, vendas_mensais,
                    reposicoes, preco_compra, preco_venda,
                    stock_minimo, tempo_reposicao, previsao_ruptura)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    tratar_nulo(row.get("produto")), tratar_nulo(row.get("categoria")),
                    tratar_nulo(row.get("quantidade_atual"), 0), tratar_nulo(row.get("vendas_mensais"), 0),
                    tratar_nulo(row.get("reposicoes"), 0), tratar_dinheiro(row.get("preco_compra")),
                    tratar_dinheiro(row.get("preco_venda")), tratar_nulo(row.get("stock_minimo"), 0),
                    tratar_nulo(row.get("tempo_reposicao"), 0), tratar_nulo(row.get("previsao_ruptura"))
                )
            )
            id_writes[linha_excel] = cursor.lastrowid


def sincronizar_financeiro_stock(cursor, df):
    print("-> A processar Financeiro do Stock...")
    df = df.dropna(subset=['id_produto'])

    for _, row in df.iterrows():
        id_val = int(row["id_produto"])
        pc = tratar_dinheiro(row.get("preco_compra"))
        pv = tratar_dinheiro(row.get("preco_venda"))
        vm = tratar_nulo(row.get("vendas_mensais"), 0)
        fm = tratar_dinheiro(row.get("faturamento_mensal"))
        cm = tratar_dinheiro(row.get("custo_mensal"))
        lm = tratar_dinheiro(row.get("lucro_mensal"))
        ml = tratar_dinheiro(row.get("margem_lucro"))
        gc = tratar_dinheiro(row.get("giro_capital"))

        cursor.execute(
            """INSERT INTO financeiro
               (id_produto, preco_compra, preco_venda, vendas_mensais, faturamento_mensal,
                custo_mensal, lucro_mensal, margem_lucro, giro_capital)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
               ON DUPLICATE KEY UPDATE
               preco_compra=%s, preco_venda=%s, vendas_mensais=%s, faturamento_mensal=%s,
               custo_mensal=%s, lucro_mensal=%s, margem_lucro=%s, giro_capital=%s""",
            (id_val, pc, pv, vm, fm, cm, lm, ml, gc,
                     pc, pv, vm, fm, cm, lm, ml, gc)
        )


# ==========================================
# MOTOR PRINCIPAL
# ==========================================
def main():
    print("A conectar a base de dados...")
    conn = get_connection()
    cursor = conn.cursor()
    arquivo_excel = "dados.xlsx"

    try:
        print("A ler o Excel (modo streaming - read_only)...")

        # ── read_only=True faz streaming do XML linha a linha ──
        # É 10-50x mais rápido que o modo normal porque NÃO carrega
        # todas as folhas em memória (incluindo guias e folhas de cálculo pesadas).
        # Apenas as folhas que acedemos são lidas.
        wb_ro = load_workbook(arquivo_excel, read_only=True, data_only=True)

        folhas_disponiveis = wb_ro.sheetnames
        sincronizadas = [f for f in FOLHAS_SYNC if f in folhas_disponiveis]
        print(f"   Folhas encontradas : {folhas_disponiveis}")
        print(f"   A sincronizar      : {sincronizadas}")

        def _ws_para_df(ws):
            """Converte uma worksheet (read_only) num DataFrame pandas."""
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return pd.DataFrame()
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            return pd.DataFrame(rows[1:], columns=headers)

        df_func      = _ws_para_df(wb_ro["funcionarios"])
        df_fin_func  = _ws_para_df(wb_ro["financeiro_funcionarios"])
        df_stock     = _ws_para_df(wb_ro["stock"])
        df_fin_stock = _ws_para_df(wb_ro["financeiro"])
        wb_ro.close()

        # ── Sincronizações SQL ──
        id_w_func  = sincronizar_funcionarios(cursor, df_func)
        id_w_stock = sincronizar_stock(cursor, df_stock)
        sincronizar_financeiro_funcionarios(cursor, df_fin_func)
        sincronizar_financeiro_stock(cursor, df_fin_stock)

        conn.commit()

        # ── Só abre o Excel para escrita se houver IDs novos para gravar ──
        # (esta é a maior optimização: load_workbook é lento)
        if id_w_func or id_w_stock:
            total_novos = len(id_w_func) + len(id_w_stock)
            print(f"   A escrever {total_novos} novo(s) ID(s) no Excel...")
            wb = load_workbook(arquivo_excel, data_only=True)
            for linha, id_val in id_w_func.items():
                wb['funcionarios'].cell(row=linha, column=1).value = id_val
            for linha, id_val in id_w_stock.items():
                wb['stock'].cell(row=linha, column=1).value = id_val
            wb.save(arquivo_excel)
            wb.close()
            print(f"   {total_novos} ID(s) escritos com sucesso.")
        else:
            print("   Sem novas linhas — Excel nao foi modificado (operacao rapida).")

        print("SINCRONIZACAO CONCLUIDA COM SUCESSO!")

    except Exception as e:
        conn.rollback()
        print(f"OCORREU UM ERRO! O banco de dados e o Excel nao foram alterados.")
        print(f"Detalhe: {e}")

    finally:
        if cursor: cursor.close()
        if conn: conn.close()


if __name__ == "__main__":
    main()