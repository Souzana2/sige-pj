import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

from conect import get_connection
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook

print("A iniciar IA: Previsão de Ruptura de Stock...")

# 1 ao 6. (Igual ao seu original)
conn = get_connection()
query = "SELECT id, quantidade_atual, vendas_mensais, reposicoes FROM stock"
df = pd.read_sql(query, conn)

# Transforma todos os espaços vazios (NaN) em 0 para a IA não crashar
df = df.fillna(0)

# Cálculo absoluto de risco (escala 0-100)
df["risco_real"] = (
    (df["vendas_mensais"] / (df["quantidade_atual"] + 1)).clip(upper=1.0) * 0.7 +
    (1 / (df["reposicoes"] + 1)) * 0.3
) * 100

X = df[["quantidade_atual", "vendas_mensais", "reposicoes"]]
y = df["risco_real"]

scaler = MinMaxScaler()
X_scaled = scaler.fit_transform(X)

model = LinearRegression()
model.fit(X_scaled, y)
predicoes = model.predict(X_scaled)

# 7. Atualizar MySQL
cursor = conn.cursor()
for produto_id, risco in zip(df["id"], predicoes):
    query = "UPDATE stock SET previsao_ruptura = %s WHERE id = %s"
    # risco já está na escala 0-100, salva diretamente
    cursor.execute(query, (float(risco), int(produto_id)))
conn.commit()
cursor.close()
conn.close()
print("✅ SQL atualizado.")

# 8. A NOVIDADE: Atualizar o Excel!
arquivo_excel = "dados.xlsx"
wb = load_workbook(arquivo_excel)
ws = wb['stock']

col_id, col_risco = None, None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "id": col_id = col
    if ws.cell(row=1, column=col).value == "previsao_ruptura": col_risco = col

if col_id and col_risco:
    dict_previsoes = {int(id_val): round(float(risco), 4) for id_val, risco in zip(df["id"], predicoes)}
    for row in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row, column=col_id).value
        if cell_id in dict_previsoes:
            ws.cell(row=row, column=col_risco).value = dict_previsoes[cell_id]

wb.save(arquivo_excel)
print("✅ Excel atualizado com as previsões de Ruptura!")