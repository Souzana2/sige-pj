# -*- coding: utf-8 -*-
import sys, io
# Força o terminal Windows a usar UTF-8 para aceitar emojis e acentos
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

"""
==============================================================
  orq.py - ORQUESTRADOR DO PIPELINE SIGE
==============================================================

  Executa na ordem correta:
    0. setup_db.py    → Garante que a BD e as tabelas existem
    1. sync.py        → Sincroniza Excel com o MySQL
    2. ml_funcionarios.py → IA: Risco de saída dos funcionários
    3. ml_stock.py    → IA: Previsão de ruptura de stock
    4. app.py         → Inicia o portal Streamlit

  Como usar:
      python orq.py

  Flags opcionais:
      python orq.py --so-sync          → corre apenas a sincronização
      python orq.py --so-ml            → corre apenas os modelos de IA
      python orq.py --so-ml-func       → corre apenas ML de funcionários
      python orq.py --so-ml-stock      → corre apenas ML de stock
      python orq.py --sem-portal       → atualiza tudo mas não abre o portal
==============================================================
"""

import sys
import time
import traceback

# ──────────────────────────────────────────────────────────────
# HELPERS DE APRESENTAÇÃO
# ──────────────────────────────────────────────────────────────
VERDE    = "\033[92m"
VERMELHO = "\033[91m"
AMARELO  = "\033[93m"
AZUL     = "\033[94m"
RESET    = "\033[0m"
NEGRITO  = "\033[1m"

def cabecalho(titulo):
    linha = "=" * 58
    print(f"\n{AZUL}{NEGRITO}{linha}")
    print(f"  {titulo}")
    print(f"{linha}{RESET}\n")

def secao(numero, titulo):
    print(f"\n{AMARELO}{NEGRITO}[PASSO {numero}] >>  {titulo}{RESET}")
    print("-" * 50)

def sucesso(msg):
    print(f"{VERDE}{NEGRITO}[OK] {msg}{RESET}")

def erro(msg):
    print(f"{VERMELHO}{NEGRITO}[ERRO] {msg}{RESET}")

def info(msg):
    print(f"   {msg}")


# ──────────────────────────────────────────────────────────────
# PASSO 0 — GARANTIA DA BASE DE DADOS (setup_db.py)
# ──────────────────────────────────────────────────────────────
def correr_setup_db():
    secao(0, "VERIFICAR / CRIAR BASE DE DADOS  (setup_db.py)")
    try:
        from setup_db import garantir_db
        ok = garantir_db()
        if ok:
            sucesso("Base de dados e tabelas verificadas/criadas com sucesso.")
        else:
            erro("Falha no setup da base de dados — verifica as credenciais.")
        return ok
    except Exception as e:
        erro(f"Erro inesperado no setup da BD!\n   Detalhe: {e}")
        traceback.print_exc()
        return False


# ──────────────────────────────────────────────────────────────
# PASSO 1 — SINCRONIZAÇÃO (sync.py)
# ──────────────────────────────────────────────────────────────
def correr_sync():
    secao(1, "SINCRONIZACAO EXCEL -> MYSQL  (sync.py)")
    try:
        from sync import main as sync_main
        sync_main()
        sucesso("Sincronizacao concluida.")
        return True
    except Exception as e:
        erro(f"Falhou a sincronizacao!\n   Detalhe: {e}")
        traceback.print_exc()
        return False


# ──────────────────────────────────────────────────────────────
# PASSO 1.5 — RECALCULO FINANCEIRO DOS FUNCIONÁRIOS
# Garante que todos os funcionarios com salario > 0 têm um
# registo correcto em financeiro_funcionarios, independentemente
# de terem sido inseridos pelo portal ou pelo Excel.
# ──────────────────────────────────────────────────────────────
def recalcular_financeiros():
    secao("1.5", "RECALCULO FINANCEIRO DOS FUNCIONARIOS")
    try:
        from conect import get_connection
        conn = get_connection()
        cursor = conn.cursor()

        # Lê todos os funcionarios activos com salario definido
        cursor.execute(
            "SELECT id, salario, horas_extra FROM funcionarios "
            "WHERE ativo = 1 AND salario > 0"
        )
        funcionarios = cursor.fetchall()

        if not funcionarios:
            info("[AVISO] Sem funcionarios activos com salario - passo ignorado.")
            conn.close()
            return True

        inseridos = 0
        atualizados = 0

        for (func_id, salario, horas_extra) in funcionarios:
            he = horas_extra or 0
            valor_hora  = salario / 160
            custo_he    = he * valor_hora * 1.5
            encargos    = round(salario * 0.2375, 2)
            custo_total = round(salario + encargos + custo_he, 2)
            turnover    = round(salario * 2, 2)

            cursor.execute(
                "SELECT custo_total_mensal FROM financeiro_funcionarios "
                "WHERE id_funcionario = %s",
                (func_id,)
            )
            row = cursor.fetchone()

            if row is None:
                # Sem registo — criar
                cursor.execute(
                    """INSERT INTO financeiro_funcionarios
                       (id_funcionario, salario, custo_horas_extra, encargos,
                        custo_total_mensal, custo_turnover_previsto)
                       VALUES (%s, %s, %s, %s, %s, %s)""",
                    (func_id, salario, custo_he, encargos, custo_total, turnover)
                )
                inseridos += 1
            elif (row[0] or 0) == 0:
                # Registo existe mas custo_total é 0 — recalcular
                cursor.execute(
                    """UPDATE financeiro_funcionarios
                       SET salario=%s, custo_horas_extra=%s, encargos=%s,
                           custo_total_mensal=%s, custo_turnover_previsto=%s
                       WHERE id_funcionario=%s""",
                    (salario, custo_he, encargos, custo_total, turnover, func_id)
                )
                atualizados += 1

        conn.commit()
        cursor.close()
        conn.close()

        if inseridos or atualizados:
            info(f"Criados: {inseridos} | Actualizados: {atualizados} registos financeiros.")
        else:
            info("Todos os registos financeiros ja estavam correctos.")

        sucesso("Financeiros dos funcionarios verificados.")
        return True

    except Exception as e:
        erro(f"Falhou o recalculo financeiro!\n   Detalhe: {e}")
        traceback.print_exc()
        return False


# ──────────────────────────────────────────────────────────────
# PASSO 2 — IA: FUNCIONÁRIOS (ml_funcionarios.py)
# ──────────────────────────────────────────────────────────────
def correr_ml_funcionarios():
    secao(2, "IA: RISCO DE SAÍDA DOS FUNCIONÁRIOS  (ml_funcionarios.py)")
    try:
        from conect import get_connection
        import pandas as pd
        from sklearn.linear_model import LinearRegression
        from sklearn.preprocessing import MinMaxScaler
        from openpyxl import load_workbook

        info("A ler dados dos funcionários da base de dados...")
        conn = get_connection()
        query = "SELECT id, assiduidade, produtividade, satisfacao, horas_extra FROM funcionarios WHERE ativo = 1"
        df = pd.read_sql(query, conn)

        if df.empty:
            info("[AVISO] Sem funcionarios activos - ML ignorado.")
            conn.close()
            return True

        # Substitui NULLs da BD por 0 (evita crash do sklearn com NaN)
        df = df.fillna(0)

        # Verifica se ha dados uteis (evita ML com tudo a zeros)
        colunas_ml = ["assiduidade", "produtividade", "satisfacao", "horas_extra"]
        if df[colunas_ml].sum().sum() == 0:
            info("[AVISO] Todos os valores de risco sao zero - ML ignorado.")
            conn.close()
            return True

        df["risco_real"] = (
            (100 - df["assiduidade"])   * 0.3 +
            (100 - df["produtividade"]) * 0.3 +
            (100 - df["satisfacao"])    * 0.3 +
            df["horas_extra"].clip(upper=100) * 0.1
        )

        X = df[["assiduidade", "produtividade", "satisfacao", "horas_extra"]]
        y = df["risco_real"]

        scaler = MinMaxScaler()
        X_scaled = scaler.fit_transform(X)

        model = LinearRegression()
        model.fit(X_scaled, y)
        predicoes = model.predict(X_scaled)

        # — Actualizar MySQL —
        info("A actualizar risco_saida no MySQL...")
        cursor = conn.cursor()
        for funcionario_id, risco in zip(df["id"], predicoes):
            cursor.execute(
                "UPDATE funcionarios SET risco_saida = %s WHERE id = %s",
                (float(risco), int(funcionario_id))
            )
        conn.commit()
        cursor.close()
        conn.close()

        # — Actualizar Excel —
        info("A actualizar risco_saida no Excel...")
        arquivo_excel = "dados.xlsx"
        wb = load_workbook(arquivo_excel)
        ws = wb["funcionarios"]

        col_id, col_risco = None, None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "id":          col_id    = col
            if val == "risco_saida": col_risco = col

        if col_id and col_risco:
            dict_prev = {
                int(id_val): round(float(risco), 4)
                for id_val, risco in zip(df["id"], predicoes)
            }
            for row in range(2, ws.max_row + 1):
                cell_id = ws.cell(row=row, column=col_id).value
                if cell_id in dict_prev:
                    ws.cell(row=row, column=col_risco).value = dict_prev[cell_id]
        else:
            info("[AVISO] Coluna 'risco_saida' nao encontrada no Excel - salvo so o SQL.")

        wb.save(arquivo_excel)
        sucesso("ML Funcionários concluído (MySQL + Excel actualizados).")
        return True

    except Exception as e:
        erro(f"Falhou o ML de Funcionários!\n   Detalhe: {e}")
        traceback.print_exc()
        return False


# ──────────────────────────────────────────────────────────────
# PASSO 3 — IA: STOCK (ml_stock.py)
# ──────────────────────────────────────────────────────────────
def correr_ml_stock():
    secao(3, "IA: PREVISÃO DE RUPTURA DE STOCK  (ml_stock.py)")
    try:
        from conect import get_connection
        import pandas as pd
        from sklearn.linear_model import LinearRegression
        from sklearn.preprocessing import MinMaxScaler
        from openpyxl import load_workbook

        info("A ler dados de stock da base de dados...")
        conn = get_connection()
        query = "SELECT id, quantidade_atual, vendas_mensais, reposicoes FROM stock"
        df = pd.read_sql(query, conn)

        if df.empty:
            info("[AVISO] Sem produtos em stock - ML ignorado.")
            conn.close()
            return True

        df = df.fillna(0)

        df["risco_real"] = (
            (df["vendas_mensais"] / (df["quantidade_atual"] + 1)).clip(upper=1.0) * 0.7 +
            (1 / (df["reposicoes"] + 1)) * 0.3
        ) * 100

        X = df[["quantidade_atual", "vendas_mensais", "reposicoes"]]
        y = df["risco_real"]

        scaler = MinMaxScaler()
        X_scaled = scaler.fit_transform(X)

        model = LinearRegression()
        model.fit(X_scaled, y)
        predicoes = model.predict(X_scaled)

        # — Actualizar MySQL —
        info("A actualizar previsao_ruptura no MySQL...")
        cursor = conn.cursor()
        for produto_id, risco in zip(df["id"], predicoes):
            cursor.execute(
                "UPDATE stock SET previsao_ruptura = %s WHERE id = %s",
                (float(risco), int(produto_id))
            )
        conn.commit()
        cursor.close()
        conn.close()

        # — Actualizar Excel —
        info("A actualizar previsao_ruptura no Excel...")
        arquivo_excel = "dados.xlsx"
        wb = load_workbook(arquivo_excel)
        ws = wb["stock"]

        col_id, col_risco = None, None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "id":                col_id    = col
            if val == "previsao_ruptura":  col_risco = col

        if col_id and col_risco:
            dict_prev = {
                int(id_val): round(float(risco), 4)
                for id_val, risco in zip(df["id"], predicoes)
            }
            for row in range(2, ws.max_row + 1):
                cell_id = ws.cell(row=row, column=col_id).value
                if cell_id in dict_prev:
                    ws.cell(row=row, column=col_risco).value = dict_prev[cell_id]
        else:
            info("[AVISO] Coluna 'previsao_ruptura' nao encontrada no Excel - salvo so o SQL.")

        wb.save(arquivo_excel)
        sucesso("ML Stock concluído (MySQL + Excel actualizados).")
        return True

    except Exception as e:
        erro(f"Falhou o ML de Stock!\n   Detalhe: {e}")
        traceback.print_exc()
        return False


# ──────────────────────────────────────────────────────────────
# PASSO 4 — INICIAR PORTAL (app.py via Streamlit)
# ──────────────────────────────────────────────────────────────
def correr_streamlit():
    secao(4, "INICIAR PORTAL DE ADMINISTRAÇÃO  (streamlit run app.py)")
    try:
        import subprocess
        info("A iniciar o portal Streamlit...")
        # Executa o Streamlit de forma síncrona/bloqueante para que o terminal fique ativo com o log
        subprocess.run([sys.executable, "-m", "streamlit", "run", "app.py"])
        return True
    except Exception as e:
        erro(f"Falhou ao iniciar o Streamlit!\n   Detalhe: {e}")
        return False


# ──────────────────────────────────────────────────────────────
# MOTOR PRINCIPAL
# ──────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]

    cabecalho("SIGE - ORQUESTRADOR DO PIPELINE  (orq.py)")
    inicio = time.time()

    # Decide que passos correr conforme as flags
    correr_s  = "--so-ml" not in args and "--so-ml-func" not in args and "--so-ml-stock" not in args
    correr_mf = "--so-sync" not in args and "--so-ml-stock" not in args
    correr_ms = "--so-sync" not in args and "--so-ml-func"  not in args
    correr_ui = "--so-sync" not in args and "--so-ml" not in args and "--so-ml-func" not in args and "--so-ml-stock" not in args and "--sem-portal" not in args

    # Flags explícitas têm prioridade
    if "--so-sync":
        if "--so-sync" in args:
            correr_s, correr_mf, correr_ms = True, False, False
    if "--so-ml" in args:
        correr_s, correr_mf, correr_ms = False, True, True
    if "--so-ml-func" in args:
        correr_s, correr_mf, correr_ms = False, True, False
    if "--so-ml-stock" in args:
        correr_s, correr_mf, correr_ms = False, False, True

    resultados = {}

    # ── Passo 0 — sempre obrigatório (garante BD) ──
    resultado_db = correr_setup_db()
    resultados["Setup BD"] = resultado_db
    if not resultado_db:
        erro("Impossivel continuar sem base de dados. Verifica se o MySQL esta a correr.")
        # Mostra resumo e sai imediatamente
        cabecalho("RESUMO FINAL  (abortado)")
        print(f"  {VERMELHO}[ERRO]{RESET}  Setup BD")
        print()
        erro("PIPELINE ABORTADO - sem ligacao a base de dados.")
        sys.exit(1)

    # ── Passo 1 ──
    if correr_s:
        resultados["Sincronizacao"] = correr_sync()
        if not resultados["Sincronizacao"]:
            erro("A sincronizacao falhou - a continuar com os dados actuais da BD.")

    # ── Passo 1.5 — sempre corre (depende so da BD, nao do Excel) ──
    resultados["Financeiros"] = recalcular_financeiros()


    # ── Passo 2 ──
    if correr_mf:
        resultados["ML Funcionarios"] = correr_ml_funcionarios()

    # ── Passo 3 ──
    if correr_ms:
        resultados["ML Stock"] = correr_ml_stock()

    # ── Sumário Final ──
    duracao = round(time.time() - inicio, 1)
    cabecalho(f"RESUMO FINAL  (concluido em {duracao}s)")

    todos_ok = True
    for nome, ok in resultados.items():
        simbolo = f"{VERDE}✅{RESET}" if ok else f"{VERMELHO}❌{RESET}"
        print(f"  {simbolo}  {nome}")
        if not ok:
            todos_ok = False

    print()
    if todos_ok:
        sucesso("PIPELINE COMPLETO - todos os sistemas actualizados!")
        if correr_ui:
            correr_streamlit()
    else:
        erro("PIPELINE TERMINADO COM ERROS - verifica os detalhes acima.")

    sys.exit(0 if todos_ok else 1)


if __name__ == "__main__":
    main()
